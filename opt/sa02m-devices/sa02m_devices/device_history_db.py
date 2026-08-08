"""Архив телеметрии ДТВ / СЭ-02м-3 (SQLite, 1 Гц, окно 30 суток).

По умолчанию пишутся все устройства из live_snapshot (dtv[] / ce[]).
PK: (ts, device_id). Путь: USB → SD → eMMC (см. stand_storage_path).
"""

from __future__ import annotations

import os
import sqlite3
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from sa02m_devices.stand_storage_path import (
    ACTIVE_NAME,
    free_bytes,
    journaling_for_fstype,
    list_history_dbs,
    mount_fstype,
    resolve_storage_target,
)

try:
    from zoneinfo import ZoneInfo

    _TZ = ZoneInfo(os.environ.get("STAND_TZ", "Europe/Moscow"))
except Exception:  # noqa: BLE001
    _TZ = timezone(timedelta(hours=3))

RETENTION_S = float(os.environ.get("STAND_DEVICES_RETENTION_S", str(30 * 86400)))
ROTATE_BYTES = int(
    os.environ.get("STAND_DEVICES_HISTORY_ROTATE_BYTES", str(3 * 1024**3))
)
ROTATE_HEADROOM = int(
    os.environ.get("STAND_DEVICES_HISTORY_ROTATE_HEADROOM", str(256 * 1024**2))
)

METRICS: dict[str, dict[str, Any]] = {
    "room_temp": {
        "table": "dtv_samples",
        "fields": ["room_temp"],
        "labels": {"room_temp": "T"},
        "label": "Температура",
        "unit": "°C",
        "device": "dtv",
    },
    "humidity": {
        "table": "dtv_samples",
        "fields": ["humidity"],
        "labels": {"humidity": "RH"},
        "label": "Влажность",
        "unit": "%",
        "device": "dtv",
    },
    "eco2_ppm": {
        "table": "dtv_samples",
        "fields": ["eco2_ppm"],
        "labels": {"eco2_ppm": "eCO₂"},
        "label": "eCO₂",
        "unit": "ppm",
        "device": "dtv",
    },
    "tvoc_mg_m3": {
        "table": "dtv_samples",
        "fields": ["tvoc_mg_m3"],
        "labels": {"tvoc_mg_m3": "TVOC"},
        "label": "TVOC",
        "unit": "mg/m³",
        "device": "dtv",
    },
    "pressure_mmhg": {
        "table": "dtv_samples",
        "fields": ["pressure_mmhg"],
        "labels": {"pressure_mmhg": "P"},
        "label": "Давление",
        "unit": "мм рт.ст.",
        "device": "dtv",
    },
    "light_pct": {
        "table": "dtv_samples",
        "fields": ["light_pct"],
        "labels": {"light_pct": "Осв."},
        "label": "Освещённость, %",
        "unit": "%",
        "device": "dtv",
    },
    "presence": {
        "table": "dtv_samples",
        "fields": ["presence"],
        "labels": {"presence": "Прис."},
        "label": "Присутствие",
        "unit": "",
        "device": "dtv",
    },
    "voltage": {
        "table": "ce_samples",
        "fields": ["voltage_a", "voltage_b", "voltage_c"],
        "labels": {"voltage_a": "Ua", "voltage_b": "Ub", "voltage_c": "Uc"},
        "label": "Напряжение Ua/Ub/Uc",
        "unit": "V",
        "device": "ce",
    },
    "current": {
        "table": "ce_samples",
        "fields": ["current_a", "current_b", "current_c"],
        "labels": {"current_a": "Ia", "current_b": "Ib", "current_c": "Ic"},
        "label": "Ток Ia/Ib/Ic",
        "unit": "A",
        "device": "ce",
    },
    "power": {
        "table": "ce_samples",
        "fields": ["power_w_a", "power_w_b", "power_w_c", "power_w_total"],
        "labels": {
            "power_w_a": "Pa",
            "power_w_b": "Pb",
            "power_w_c": "Pc",
            "power_w_total": "P∑",
        },
        "label": "Мощность",
        "unit": "W",
        "device": "ce",
    },
    "frequency_hz": {
        "table": "ce_samples",
        "fields": ["frequency_hz"],
        "labels": {"frequency_hz": "f"},
        "label": "Частота",
        "unit": "Hz",
        "device": "ce",
    },
    "energy_kwh_import": {
        "table": "ce_samples",
        "fields": ["energy_kwh_import"],
        "labels": {"energy_kwh_import": "E"},
        "label": "Энергия (импорт)",
        "unit": "kWh",
        "device": "ce",
    },
}

HISTORY_GROUPS: dict[str, list[str]] = {
    "climate": [
        "room_temp", "humidity", "eco2_ppm", "tvoc_mg_m3",
        "pressure_mmhg", "light_pct", "presence",
    ],
    "energy": [
        "voltage", "current", "power", "frequency_hz", "energy_kwh_import",
    ],
}

# range_key → (window_s OR None, chart_bucket_s). None window = календарный режим.
RANGES: dict[str, tuple[float | None, float]] = {
    "1h": (3600.0, 5.0),
    "6h": (6 * 3600.0, 30.0),
    "24h": (24 * 3600.0, 120.0),
    "7d": (7 * 86400.0, 600.0),
    "30d": (30 * 86400.0, 3600.0),
    "mtd": (None, 3600.0),  # с начала месяца
    "month": (None, 86400.0),  # предыдущий календарный месяц
}

# Агрегация для текстового экспорта
EXPORT_BUCKET_S: dict[str, float] = {
    "1h": 60.0,
    "6h": 1800.0,
    "24h": 3600.0,
    "7d": 86400.0,
    "30d": 86400.0,
    "mtd": 86400.0,
    "month": 86400.0,
}

RANGE_LABELS_RU: dict[str, str] = {
    "1h": "1 ч",
    "6h": "6 ч",
    "24h": "24 ч",
    "7d": "7 д",
    "30d": "30 д",
    "mtd": "с начала месяца",
    "month": "за месяц",
}

# Ориентир для ЮЛ г. Москва (1 ц.к., с НДС) — пользователь правит в UI.
DEFAULT_KWH_RUB = float(os.environ.get("STAND_DEVICES_KWH_RUB", "10.50"))


def _now_local() -> datetime:
    return datetime.now(_TZ)


def resolve_time_range(range_key: str) -> tuple[float, float, float]:
    """Вернуть (t0, t1, chart_bucket_s) для range_key."""
    key = range_key if range_key in RANGES else "1h"
    window_s, bucket_s = RANGES[key]
    now = time.time()
    if window_s is not None:
        return now - float(window_s), now, float(bucket_s)
    local = _now_local()
    if key == "mtd":
        start = local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return start.timestamp(), now, float(bucket_s)
    # previous calendar month
    first_this = local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_prev = first_this - timedelta(seconds=1)
    start_prev = last_prev.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end_prev = first_this
    return start_prev.timestamp(), end_prev.timestamp(), float(bucket_s)


def export_bucket_s(range_key: str) -> float:
    return float(EXPORT_BUCKET_S.get(range_key) or EXPORT_BUCKET_S["1h"])

_CREATE_DTV = """
    CREATE TABLE IF NOT EXISTS dtv_samples (
        ts REAL NOT NULL,
        device_id TEXT NOT NULL DEFAULT '',
        room_temp REAL,
        humidity REAL,
        eco2_ppm REAL,
        tvoc_mg_m3 REAL,
        pressure_mmhg REAL,
        light_pct REAL,
        presence REAL,
        PRIMARY KEY (ts, device_id)
    );
"""
_CREATE_CE = """
    CREATE TABLE IF NOT EXISTS ce_samples (
        ts REAL NOT NULL,
        device_id TEXT NOT NULL DEFAULT '',
        voltage_a REAL,
        voltage_b REAL,
        voltage_c REAL,
        current_a REAL,
        current_b REAL,
        current_c REAL,
        power_w_a REAL,
        power_w_b REAL,
        power_w_c REAL,
        power_w_total REAL,
        frequency_hz REAL,
        energy_kwh_import REAL,
        PRIMARY KEY (ts, device_id)
    );
"""


def _ensure_ce_power_phase_cols(conn: sqlite3.Connection) -> None:
    cols = {
        str(r[1])
        for r in conn.execute("PRAGMA table_info(ce_samples)").fetchall()
    }
    for col in ("power_w_a", "power_w_b", "power_w_c"):
        if col not in cols:
            conn.execute(f"ALTER TABLE ce_samples ADD COLUMN {col} REAL")


def db_path(path: Path | None = None) -> Path:
    if path is not None:
        return Path(path)
    forced = str(os.environ.get("STAND_DEVICES_HISTORY_DB") or "").strip()
    if forced:
        return Path(forced)
    return resolve_storage_target().active_path


def storage_status() -> dict[str, Any]:
    """Метаданные носителя для API."""
    forced = str(os.environ.get("STAND_DEVICES_HISTORY_DB") or "").strip()
    if forced:
        p = Path(forced)
        return {
            "history_backend": "force",
            "history_mount": str(p.parent),
            "history_db_dir": str(p.parent),
            "history_db_path": str(p),
            "history_free_bytes": free_bytes(p.parent),
            "history_fstype": mount_fstype(p.parent),
            "history_archives_count": len(
                list(p.parent.glob("devices_history_*.db"))
            ),
        }
    return resolve_storage_target(force_refresh=True).as_dict()


def _as_device_list(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [d for d in value if isinstance(d, dict)]
    if isinstance(value, dict):
        return [value]
    return []


def _needs_pk_migration(conn: sqlite3.Connection, table: str) -> bool:
    row = conn.execute(
        "SELECT sql FROM sqlite_master WHERE type='table' AND name=?",
        (table,),
    ).fetchone()
    if not row or not row[0]:
        return False
    sql = row[0].upper().replace(" ", "")
    if "PRIMARYKEY(TS,DEVICE_ID)" in sql:
        return False
    if "TSPREALPRIMARYKEY" in sql or "TSREALPRIMARYKEY" in sql:
        return True
    return "PRIMARYKEY(TS," not in sql and "PRIMARY KEY" in row[0].upper()


def _migrate_table(conn: sqlite3.Connection, table: str, create_sql: str) -> None:
    tmp = f"{table}__mig"
    conn.executescript(
        f"""
        DROP TABLE IF EXISTS {tmp};
        {create_sql.replace(table, tmp, 1)}
        INSERT OR IGNORE INTO {tmp} SELECT * FROM {table};
        DROP TABLE {table};
        ALTER TABLE {tmp} RENAME TO {table};
        """
    )


def ensure_schema(path: Path | None = None) -> Path:
    """Создать файл БД и схему; вернуть путь."""
    p = db_path(path)
    _connect(p).close()
    return p


def _connect(path: Path | None = None) -> sqlite3.Connection:
    p = db_path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(p), timeout=30.0)
    fst = mount_fstype(p.parent)
    journal, sync = journaling_for_fstype(fst)
    conn.execute(f"PRAGMA journal_mode={journal}")
    conn.execute(f"PRAGMA synchronous={sync}")
    conn.executescript(_CREATE_DTV + _CREATE_CE)
    with conn:
        if _needs_pk_migration(conn, "dtv_samples"):
            _migrate_table(conn, "dtv_samples", _CREATE_DTV)
        if _needs_pk_migration(conn, "ce_samples"):
            _migrate_table(conn, "ce_samples", _CREATE_CE)
        _ensure_ce_power_phase_cols(conn)
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_dtv_device_ts "
            "ON dtv_samples(device_id, ts)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_ce_device_ts "
            "ON ce_samples(device_id, ts)"
        )
    return conn


def _insert_dtv(conn: sqlite3.Connection, ts: float, dtv: dict[str, Any]) -> None:
    if not (
        dtv.get("ok")
        or any(
            dtv.get(k) is not None
            for k in (
                "room_temp", "humidity", "eco2_ppm", "tvoc_mg_m3",
                "pressure_mmhg", "light_pct",
            )
        )
    ):
        return
    conn.execute(
        """
        INSERT OR REPLACE INTO dtv_samples(
            ts, device_id, room_temp, humidity, eco2_ppm,
            tvoc_mg_m3, pressure_mmhg, light_pct, presence
        ) VALUES (?,?,?,?,?,?,?,?,?)
        """,
        (
            ts,
            str(dtv.get("id") or ""),
            dtv.get("room_temp"),
            dtv.get("humidity"),
            dtv.get("eco2_ppm"),
            dtv.get("tvoc_mg_m3"),
            dtv.get("pressure_mmhg"),
            dtv.get("light_pct"),
            dtv.get("presence"),
        ),
    )


def _insert_ce(conn: sqlite3.Connection, ts: float, ce: dict[str, Any]) -> None:
    volt = ce.get("voltage") if isinstance(ce.get("voltage"), dict) else {}
    curr = ce.get("current") if isinstance(ce.get("current"), dict) else {}
    pwr = ce.get("power_w") if isinstance(ce.get("power_w"), dict) else {}
    if not (ce.get("ok") or volt.get("a") is not None):
        return
    conn.execute(
        """
        INSERT OR REPLACE INTO ce_samples(
            ts, device_id, voltage_a, voltage_b, voltage_c,
            current_a, current_b, current_c,
            power_w_a, power_w_b, power_w_c, power_w_total,
            frequency_hz, energy_kwh_import
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            ts,
            str(ce.get("id") or ""),
            volt.get("a"), volt.get("b"), volt.get("c"),
            curr.get("a"), curr.get("b"), curr.get("c"),
            pwr.get("a"), pwr.get("b"), pwr.get("c"),
            pwr.get("total"),
            ce.get("frequency_hz"),
            ce.get("energy_kwh_import"),
        ),
    )


def insert_sample(snapshot: dict[str, Any], path: Path | None = None) -> None:
    """Записать снимок live_snapshot() — все ДТВ/СЭ в списках (по умолчанию)."""
    ts = float(snapshot.get("ts") or time.time())
    dtv_list = _as_device_list(snapshot.get("dtv"))
    ce_list = _as_device_list(snapshot.get("ce"))
    conn = _connect(path)
    try:
        with conn:
            for dtv in dtv_list:
                _insert_dtv(conn, ts, dtv)
            for ce in ce_list:
                _insert_ce(conn, ts, ce)
    finally:
        conn.close()


def rotate_if_needed(path: Path | None = None) -> dict[str, Any]:
    """При размере active ≥ 3 ГиБ — архивировать и создать новый active."""
    p = db_path(path)
    if not p.is_file():
        return {"rotated": False, "reason": "no_file"}
    try:
        size = p.stat().st_size
    except OSError as exc:
        return {"rotated": False, "reason": str(exc)}
    if size < ROTATE_BYTES:
        return {"rotated": False, "size": size, "threshold": ROTATE_BYTES}
    free = free_bytes(p.parent)
    if free < ROTATE_HEADROOM:
        return {
            "rotated": False,
            "skipped": "low_space",
            "size": size,
            "free_bytes": free,
            "headroom": ROTATE_HEADROOM,
        }
    # checkpoint + rename
    try:
        conn = sqlite3.connect(str(p), timeout=60.0)
        try:
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            conn.commit()
        finally:
            conn.close()
    except sqlite3.Error:
        pass
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    archive = p.parent / f"devices_history_{stamp}.db"
    n = 0
    while archive.exists():
        n += 1
        archive = p.parent / f"devices_history_{stamp}_{n}.db"
    p.replace(archive)
    for suf in ("-wal", "-shm"):
        side = Path(str(p) + suf)
        if side.is_file():
            try:
                side.unlink()
            except OSError:
                pass
    ensure_schema(p)
    return {
        "rotated": True,
        "archive": str(archive),
        "active": str(p),
        "size": size,
    }


def purge_old(path: Path | None = None, *, now: float | None = None) -> dict[str, int]:
    cutoff = float(now if now is not None else time.time()) - RETENTION_S
    conn = _connect(path)
    try:
        with conn:
            c1 = conn.execute(
                "DELETE FROM dtv_samples WHERE ts < ?", (cutoff,)
            ).rowcount
            c2 = conn.execute(
                "DELETE FROM ce_samples WHERE ts < ?", (cutoff,)
            ).rowcount
        ev_deleted = 0
        try:
            from sa02m_devices.device_events import purge_events

            ev_deleted = purge_events(path=path, cutoff=cutoff)
        except Exception:  # noqa: BLE001
            ev_deleted = 0
        return {
            "dtv_deleted": int(c1 or 0),
            "ce_deleted": int(c2 or 0),
            "events_deleted": int(ev_deleted),
            "cutoff": cutoff,
        }
    finally:
        conn.close()


def _query_series(
    conn: sqlite3.Connection,
    table: str,
    fields: list[str],
    labels: dict[str, str],
    t0: float,
    t1: float,
    bucket_s: float,
    device_id: str | None = None,
    *,
    agg: str = "avg",
) -> list[dict[str, Any]]:
    cols = ", ".join(fields)
    where = " WHERE ts >= ? AND ts <= ?"
    params: list[Any] = [t0, t1]
    if device_id:
        where += " AND device_id = ?"
        params.append(device_id)
    agg_fn = "max" if agg == "max" else "avg"
    if bucket_s <= 1.0:
        sql = f"SELECT ts, {cols} FROM {table}{where} ORDER BY ts"
        rows = conn.execute(sql, params).fetchall()
        series_map: dict[str, list[list[Any]]] = {f: [] for f in fields}
        for row in rows:
            ts_ms = int(float(row[0]) * 1000)
            for i, field in enumerate(fields):
                val = row[i + 1]
                if val is None:
                    continue
                try:
                    series_map[field].append([ts_ms, float(val)])
                except (TypeError, ValueError):
                    continue
    else:
        # По одному полю: бакеты с NULL в соседних колонках не теряют точки
        series_map = {f: [] for f in fields}
        for field in fields:
            sql = (
                f"SELECT cast(ts / ? as integer) * ? AS bucket, {agg_fn}({field})"
                f" FROM {table}{where} AND {field} IS NOT NULL"
                f" GROUP BY bucket ORDER BY bucket"
            )
            try:
                rows = conn.execute(sql, [bucket_s, bucket_s, *params]).fetchall()
            except sqlite3.Error:
                continue
            for row in rows:
                if row[1] is None:
                    continue
                try:
                    series_map[field].append(
                        [int(float(row[0]) * 1000), float(row[1])]
                    )
                except (TypeError, ValueError):
                    continue
    return [
        {
            "field": f,
            "label": labels.get(f, f),
            "points": series_map[f],
        }
        for f in fields
        if series_map[f]
    ]


def _merge_series_lists(
    parts: list[list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    """Объединить series из нескольких БД по field, точки по ts."""
    by_field: dict[str, dict[str, Any]] = {}
    for series_list in parts:
        for ser in series_list:
            field = ser["field"]
            if field not in by_field:
                by_field[field] = {
                    "field": field,
                    "label": ser.get("label", field),
                    "points": [],
                }
            by_field[field]["points"].extend(ser.get("points") or [])
    out: list[dict[str, Any]] = []
    for field, ser in by_field.items():
        pts = ser["points"]
        # dedupe by ts keep last
        seen: dict[int, float] = {}
        for ts_ms, val in pts:
            seen[int(ts_ms)] = val
        merged = [[ts, seen[ts]] for ts in sorted(seen)]
        if merged:
            out.append({
                "field": field,
                "label": ser["label"],
                "points": merged,
            })
    return out


def _read_paths(path: Path | None) -> list[Path]:
    if path is not None:
        return [Path(path)]
    forced = str(os.environ.get("STAND_DEVICES_HISTORY_DB") or "").strip()
    if forced:
        return [Path(forced)]
    target = resolve_storage_target()
    paths = list_history_dbs(target.directory)
    return paths if paths else [target.active_path]


def history(
    metric_id: str,
    range_key: str = "1h",
    path: Path | None = None,
    *,
    device_id: str | None = None,
    bucket_s: float | None = None,
    agg: str = "avg",
) -> dict[str, Any]:
    meta = METRICS.get(metric_id)
    if not meta:
        return {"ok": False, "error": "unknown metric", "metric": metric_id}
    if range_key not in RANGES:
        range_key = "1h"
    t0, t1, chart_bucket = resolve_time_range(range_key)
    if bucket_s is None:
        bucket_s = chart_bucket
    did = (device_id or "").strip() or None
    parts: list[list[dict[str, Any]]] = []
    for dbfile in _read_paths(path):
        if not dbfile.is_file():
            continue
        try:
            conn = _connect(dbfile)
        except sqlite3.Error:
            continue
        try:
            if not did:
                row = conn.execute(
                    f"SELECT device_id FROM {meta['table']}"
                    f" WHERE ts >= ? AND ts <= ? AND device_id != ''"
                    f" ORDER BY device_id LIMIT 1",
                    (t0, t1),
                ).fetchone()
                if row:
                    did = str(row[0])
            parts.append(
                _query_series(
                    conn,
                    meta["table"],
                    list(meta["fields"]),
                    dict(meta["labels"]),
                    t0,
                    t1,
                    bucket_s,
                    device_id=did,
                    agg=agg,
                )
            )
        finally:
            conn.close()
    series = _merge_series_lists(parts)
    status = storage_status() if path is None else {}
    return {
        "ok": True,
        "metric": metric_id,
        "label": meta["label"],
        "unit": meta["unit"],
        "device": meta["device"],
        "device_id": did or "",
        "range": range_key,
        "t0": t0,
        "t1": t1,
        "t0_ms": int(t0 * 1000),
        "t1_ms": int(t1 * 1000),
        "series": series,
        **status,
    }


def history_batch(
    range_key: str = "1h",
    group: str | None = None,
    metric_ids: list[str] | None = None,
    path: Path | None = None,
    *,
    device_id: str | None = None,
) -> dict[str, Any]:
    if range_key not in RANGES:
        range_key = "1h"
    if metric_ids:
        ids = [m for m in metric_ids if m in METRICS]
    elif group and group in HISTORY_GROUPS:
        ids = list(HISTORY_GROUPS[group])
    else:
        return {
            "ok": False,
            "error": "specify group=climate|energy or metrics=…",
        }
    out_metrics: list[dict[str, Any]] = []
    errors: list[str] = []
    for mid in ids:
        one = history(mid, range_key, path=path, device_id=device_id)
        if not one.get("ok"):
            errors.append(f"{mid}: {one.get('error', 'fail')}")
            out_metrics.append({
                "metric": mid,
                "label": METRICS[mid]["label"],
                "unit": METRICS[mid]["unit"],
                "device": METRICS[mid]["device"],
                "device_id": device_id or "",
                "series": [],
                "error": one.get("error"),
            })
            continue
        out_metrics.append({
            "metric": mid,
            "label": one["label"],
            "unit": one["unit"],
            "device": one["device"],
            "device_id": one.get("device_id") or "",
            "series": one.get("series") or [],
        })
    status = storage_status() if path is None else {}
    t0, t1, _b = resolve_time_range(range_key)
    return {
        "ok": True,
        "range": range_key,
        "group": group or "custom",
        "device_id": (device_id or "").strip(),
        "t0": t0,
        "t1": t1,
        "t0_ms": int(t0 * 1000),
        "t1_ms": int(t1 * 1000),
        "metrics": out_metrics,
        "errors": errors,
        **status,
    }


def _number_is_finite(v: Any) -> bool:
    try:
        f = float(v)
    except (TypeError, ValueError):
        return False
    return f == f and abs(f) != float("inf")


def period_summary_ce(
    range_key: str = "1h",
    path: Path | None = None,
    *,
    device_id: str | None = None,
    kwh_rub: float | None = None,
) -> dict[str, Any]:
    """Сводка СЭ за период: ср. мощность по фазам, ΔE, стоимость."""
    if range_key not in RANGES:
        range_key = "1h"
    t0, t1, _bucket = resolve_time_range(range_key)
    did = (device_id or "").strip() or None
    tariff = float(kwh_rub if kwh_rub is not None else DEFAULT_KWH_RUB)

    sum_pa = sum_pb = sum_pc = sum_pt = 0.0
    n_pa = n_pb = n_pc = n_pt = 0
    e_first: float | None = None
    e_last: float | None = None
    e_first_ts: float | None = None
    e_last_ts: float | None = None
    samples = 0

    for dbfile in _read_paths(path):
        if not dbfile.is_file():
            continue
        try:
            conn = _connect(dbfile)
        except sqlite3.Error:
            continue
        try:
            if not did:
                row = conn.execute(
                    "SELECT device_id FROM ce_samples"
                    " WHERE ts >= ? AND ts <= ? AND device_id != ''"
                    " ORDER BY device_id LIMIT 1",
                    (t0, t1),
                ).fetchone()
                if row:
                    did = str(row[0])
            if not did:
                continue
            where = " WHERE ts >= ? AND ts <= ? AND device_id = ?"
            params: list[Any] = [t0, t1, did]
            # averages
            row = conn.execute(
                "SELECT"
                " avg(power_w_a), count(power_w_a),"
                " avg(power_w_b), count(power_w_b),"
                " avg(power_w_c), count(power_w_c),"
                " avg(power_w_total), count(power_w_total),"
                " count(*)"
                f" FROM ce_samples{where}",
                params,
            ).fetchone()
            if row and int(row[8] or 0) > 0:
                samples += int(row[8])
                if row[0] is not None and int(row[1] or 0) > 0:
                    sum_pa += float(row[0]) * int(row[1])
                    n_pa += int(row[1])
                if row[2] is not None and int(row[3] or 0) > 0:
                    sum_pb += float(row[2]) * int(row[3])
                    n_pb += int(row[3])
                if row[4] is not None and int(row[5] or 0) > 0:
                    sum_pc += float(row[4]) * int(row[5])
                    n_pc += int(row[5])
                if row[6] is not None and int(row[7] or 0) > 0:
                    sum_pt += float(row[6]) * int(row[7])
                    n_pt += int(row[7])
            # energy endpoints
            row0 = conn.execute(
                "SELECT ts, energy_kwh_import FROM ce_samples"
                f"{where} AND energy_kwh_import IS NOT NULL"
                " ORDER BY ts ASC LIMIT 1",
                params,
            ).fetchone()
            row1 = conn.execute(
                "SELECT ts, energy_kwh_import FROM ce_samples"
                f"{where} AND energy_kwh_import IS NOT NULL"
                " ORDER BY ts DESC LIMIT 1",
                params,
            ).fetchone()
            if row0 and _number_is_finite(row0[1]):
                ts0, e0 = float(row0[0]), float(row0[1])
                if e_first is None or ts0 < (e_first_ts or ts0):
                    e_first, e_first_ts = e0, ts0
            if row1 and _number_is_finite(row1[1]):
                ts1, e1 = float(row1[0]), float(row1[1])
                if e_last is None or ts1 > (e_last_ts or ts1):
                    e_last, e_last_ts = e1, ts1
        finally:
            conn.close()

    def _avg(s: float, n: int) -> float | None:
        return None if n <= 0 else round(s / n, 3)

    # Стоимость только по энергии (кВт·ч), не по средней мощности (кВт/Вт):
    # cost_rub = ΔE_kWh × ₽/кВт·ч. Мощность в ответе — справочно.
    energy_kwh = None
    if e_first is not None and e_last is not None:
        energy_kwh = round(max(0.0, e_last - e_first), 4)
    cost = None if energy_kwh is None else round(float(energy_kwh) * tariff, 2)

    return {
        "ok": True,
        "device": "ce",
        "device_id": did or "",
        "range": range_key,
        "t0": t0,
        "t1": t1,
        "samples": samples,
        "power_w": {
            "a": _avg(sum_pa, n_pa),
            "b": _avg(sum_pb, n_pb),
            "c": _avg(sum_pc, n_pc),
            "total": _avg(sum_pt, n_pt),
        },
        "energy_kwh_import": {
            "first": e_first,
            "last": e_last,
            "delta": energy_kwh,
            "unit": "kWh",
        },
        "kwh_rub": tariff,
        "kwh_rub_default": DEFAULT_KWH_RUB,
        "kwh_rub_unit": "RUB/kWh",
        "cost_rub": cost,
        "cost_basis": "energy_kwh",
        "cost_note": "Стоимость = ΔE (кВт·ч) × тариф (₽/кВт·ч)",
        **(storage_status() if path is None else {}),
    }


def _fmt_export_ts(ts_s: float, bucket_s: float) -> str:
    dt = datetime.fromtimestamp(float(ts_s), tz=_TZ)
    if bucket_s >= 86400:
        return dt.strftime("%Y-%m-%d")
    if bucket_s >= 3600:
        return dt.strftime("%Y-%m-%d %H:00")
    return dt.strftime("%Y-%m-%d %H:%M")


def _export_bucket_label(bucket_s: float) -> str:
    if bucket_s <= 60:
        return "1 мин"
    if bucket_s <= 1800:
        return "30 мин"
    if bucket_s <= 3600:
        return "1 ч"
    return "1 сут"


def _export_col_title(metric_id: str, field: str) -> str:
    meta = METRICS[metric_id]
    lab = str(meta["labels"].get(field) or field)
    unit = str(meta.get("unit") or "").strip()
    return f"{lab}, {unit}" if unit else lab


def collect_export_table(
    range_key: str = "1h",
    *,
    metric_id: str | None = None,
    group: str | None = None,
    device_id: str | None = None,
    path: Path | None = None,
) -> dict[str, Any]:
    """Собрать одну таблицу (время × колонки) для экспорта."""
    if range_key not in RANGES:
        range_key = "1h"
    bucket = export_bucket_s(range_key)
    t0, t1, _ = resolve_time_range(range_key)
    did = (device_id or "").strip()

    if group and group in HISTORY_GROUPS:
        metric_ids = list(HISTORY_GROUPS[group])
    elif metric_id and metric_id in METRICS:
        metric_ids = [metric_id]
    else:
        return {
            "ok": False,
            "error": "укажите metric=… или group=climate|energy",
            "headers": ["Время"],
            "rows": [],
            "device_id": did,
            "range": range_key,
            "bucket_s": bucket,
            "t0": t0,
            "t1": t1,
            "metric_ids": [],
            "kind": "device",
        }

    col_fields: list[tuple[str, str]] = []  # (metric_id, field)
    for mid in metric_ids:
        for field in METRICS[mid]["fields"]:
            col_fields.append((mid, field))

    by_ts: dict[int, dict[str, float]] = {}
    for mid in metric_ids:
        one = history(
            mid,
            range_key,
            path=path,
            device_id=did or None,
            bucket_s=bucket,
            agg="max" if mid == "energy_kwh_import" else "avg",
        )
        if not did and one.get("device_id"):
            did = str(one["device_id"])
        for ser in one.get("series") or []:
            field = str(ser.get("field") or "")
            if not field:
                continue
            key = f"{mid}:{field}"
            for ts_ms, val in ser.get("points") or []:
                try:
                    by_ts.setdefault(int(ts_ms), {})[key] = float(val)
                except (TypeError, ValueError):
                    continue

    headers = ["Время"] + [_export_col_title(m, f) for m, f in col_fields]
    rows: list[list[Any]] = []
    for ts_ms in sorted(by_ts):
        cells: list[Any] = [_fmt_export_ts(ts_ms / 1000.0, bucket)]
        vals = by_ts[ts_ms]
        for mid, field in col_fields:
            v = vals.get(f"{mid}:{field}")
            cells.append(None if v is None else float(v))
        rows.append(cells)

    kind = "dtv"
    if metric_ids and all(METRICS[m]["device"] == "ce" for m in metric_ids):
        kind = "ce"
    elif metric_ids and any(METRICS[m]["device"] == "ce" for m in metric_ids):
        kind = "mixed"

    return {
        "ok": True,
        "error": "",
        "headers": headers,
        "rows": rows,
        "device_id": did,
        "range": range_key,
        "bucket_s": bucket,
        "t0": t0,
        "t1": t1,
        "metric_ids": metric_ids,
        "kind": kind,
        "title": (
            METRICS[metric_ids[0]]["label"]
            if len(metric_ids) == 1
            else ("Климат" if group == "climate" else "Энергия" if group == "energy" else "Данные")
        ),
    }


def _xml_escape(s: str) -> str:
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _export_xlsx_minimal(table: dict[str, Any], filename: str) -> bytes:
    """OOXML .xlsx без openpyxl (достаточно для Excel/LibreOffice)."""
    import zipfile
    from io import BytesIO

    def cell_ref(row: int, col: int) -> str:
        n = col
        letters = ""
        while n:
            n, rem = divmod(n - 1, 26)
            letters = chr(65 + rem) + letters
        return f"{letters}{row}"

    def inline_cell(row: int, col: int, val: Any) -> str:
        ref = cell_ref(row, col)
        if val is None:
            return f'<c r="{ref}"/>'
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            return f'<c r="{ref}"><v>{val}</v></c>'
        text = _xml_escape(str(val))
        return (
            f'<c r="{ref}" t="inlineStr"><is><t>{text}</t></is></c>'
        )

    meta_rows = [
        ("Устройство", table.get("device_id") or "—"),
        ("Метрика", table.get("title") or "—"),
        ("Строк", str(len(table.get("rows") or []))),
    ]
    headers = list(table.get("headers") or ["Время"])
    data_rows = list(table.get("rows") or [])
    sheet_rows: list[str] = []
    r = 1
    for k, v in meta_rows:
        sheet_rows.append(
            f'<row r="{r}">'
            f"{inline_cell(r, 1, k)}{inline_cell(r, 2, v)}"
            f"</row>"
        )
        r += 1
    r = 5
    sheet_rows.append(
        f'<row r="{r}">'
        + "".join(inline_cell(r, c, h) for c, h in enumerate(headers, start=1))
        + "</row>"
    )
    if not data_rows:
        r = 6
        sheet_rows.append(
            f'<row r="{r}">{inline_cell(r, 1, "(нет точек за период)")}</row>'
        )
    else:
        for row in data_rows:
            r += 1
            sheet_rows.append(
                f'<row r="{r}">'
                + "".join(
                    inline_cell(r, c, val) for c, val in enumerate(row, start=1)
                )
                + "</row>"
            )

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(sheet_rows)}</sheetData></worksheet>"
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    wb_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Данные" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    wb_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("xl/workbook.xml", wb_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    _ = filename  # kept for call-site symmetry
    return buf.getvalue()


def export_xlsx(
    range_key: str = "1h",
    *,
    metric_id: str | None = None,
    group: str | None = None,
    device_id: str | None = None,
    path: Path | None = None,
) -> tuple[bytes, str]:
    """Выгрузка в Excel (.xlsx) с таблицей. Возвращает (bytes, filename)."""
    from io import BytesIO

    table = collect_export_table(
        range_key, metric_id=metric_id, group=group, device_id=device_id, path=path
    )
    stamp = _now_local().strftime("%Y%m%d_%H%M%S")
    mid_part = metric_id or group or "data"
    safe_id = (str(table.get("device_id") or "device")).replace("/", "-")
    kind = str(table.get("kind") or "device")
    filename = f"{kind}_export_{safe_id}_{mid_part}_{range_key}_{stamp}.xlsx"

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return _export_xlsx_minimal(table, filename), filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Данные"

    header_font = Font(bold=True)
    meta_font = Font(bold=True, color="334455")
    meta_fill = PatternFill("solid", fgColor="E8EEF5")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    meta_rows = [
        ("Устройство", table.get("device_id") or "—"),
        ("Метрика", table.get("title") or mid_part),
        (
            "Период",
            f"{RANGE_LABELS_RU.get(range_key, range_key)} "
            f"({_fmt_export_ts(float(table['t0']), 60)} — "
            f"{_fmt_export_ts(float(table['t1']), 60)})",
        ),
        ("Шаг", _export_bucket_label(float(table["bucket_s"]))),
        ("Строк", len(table.get("rows") or [])),
    ]
    for i, (k, v) in enumerate(meta_rows, start=1):
        ws.cell(i, 1, k).font = meta_font
        ws.cell(i, 1).fill = meta_fill
        ws.cell(i, 2, v)

    header_row = 7
    headers = list(table.get("headers") or ["Время"])
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, h)
        cell.font = header_font
        cell.alignment = center

    data_rows = list(table.get("rows") or [])
    if not data_rows:
        ws.cell(header_row + 1, 1, "(нет точек за период)")
    else:
        for r_i, row in enumerate(data_rows, start=header_row + 1):
            for c_i, val in enumerate(row, start=1):
                cell = ws.cell(r_i, c_i, val)
                cell.alignment = center
                if c_i > 1 and isinstance(val, float):
                    cell.number_format = "0.###"

    last_data_row = header_row + max(1, len(data_rows))
    last_col = max(1, len(headers))
    # Excel Table (полноценная таблица)
    if data_rows:
        ref = f"A{header_row}:{get_column_letter(last_col)}{last_data_row}"
        tab = Table(displayName="ExportData", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)

    ws.column_dimensions["A"].width = 20
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.freeze_panes = f"A{header_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename


def export_text(
    range_key: str = "1h",
    *,
    metric_id: str | None = None,
    group: str | None = None,
    device_id: str | None = None,
    path: Path | None = None,
) -> tuple[str, str]:
    """Текстовая выгрузка (TSV) — совместимость; основной формат: export_xlsx."""
    table = collect_export_table(
        range_key, metric_id=metric_id, group=group, device_id=device_id, path=path
    )
    if not table.get("ok"):
        return f"# error: {table.get('error')}\n", "export_error.txt"
    headers = table["headers"]
    lines = [
        f"# устройство: {table.get('device_id') or '—'}",
        f"# метрика: {table.get('title')}",
        f"# период: {RANGE_LABELS_RU.get(range_key, range_key)}"
        f" ({_fmt_export_ts(float(table['t0']), 60)} — "
        f"{_fmt_export_ts(float(table['t1']), 60)})",
        f"# шаг: {_export_bucket_label(float(table['bucket_s']))}",
        "\t".join(str(h) for h in headers),
    ]
    for row in table.get("rows") or []:
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            elif isinstance(v, float):
                cells.append(f"{v:.6g}")
            else:
                cells.append(str(v))
        lines.append("\t".join(cells))
    if not table.get("rows"):
        lines.append("# (нет точек)")
    body = "\n".join(lines) + "\n"
    stamp = _now_local().strftime("%Y%m%d_%H%M%S")
    mid_part = metric_id or group or "data"
    safe_id = (str(table.get("device_id") or "device")).replace("/", "-")
    kind = str(table.get("kind") or "device")
    filename = f"{kind}_export_{safe_id}_{mid_part}_{range_key}_{stamp}.txt"
    return body, filename


# совместимость: старое имя константы
DEFAULT_DB_PATH = Path("/var/lib/sa02m-stand") / ACTIVE_NAME
